import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

xlsx_path = r"D:/06.3지국 전용방/통합부재법전_2026-05-05.xlsx"

def main():
    print(f"[{os.path.basename(xlsx_path)}] 통합부재법전 검사 시작...")
    if not os.path.exists(xlsx_path):
        print(f"오류: 파일 없음: {xlsx_path}")
        return
        
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print("시트 목록:", wb.sheetnames)
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"\n--- 시트: {sheetname} (최대 {sheet.max_row}행) ---")
        # 처음 15행 읽기
        rows = list(sheet.iter_rows(max_row=15, values_only=True))
        for i, row in enumerate(rows, 1):
            # None이 아닌 값들만 간소하게 출력
            clean_row = [str(val)[:30] if val is not None else "" for val in row]
            # 비어있는 행 제외
            if any(clean_row):
                print(f"  Row {i:2d}: {clean_row}")

if __name__ == "__main__":
    main()
