"""
벽체 일람표 OLE 추출 v2 — 동 분류 정정판

검수 결과 (2026-05-04, 이천):
- 한 xlsx = 한 동(BarList) 또는 다중 동(시트명이 NNN동)
- 동 표제는 BarList 시트 상단 <NNN동> 형식 1개만
- 관대 패턴 NNN동 검출은 0건이라 엄격 패턴으로 충분
- 시트명 분리형 61개 별도 처리

소스: tmp_analysis/ole_all/*.xlsx (981개)
적재: jiguk3_assets.db members 테이블 (region=DONG{nnn}_WALL)
"""

import os, re, sys, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from collections import Counter
from core import asset_db

XLSX_FILES = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
print(f"[입력] xlsx {len(XLSX_FILES)}개")

DONG_HEADER_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')
DONG_SHEET_RE = re.compile(r'^(\d{3})\s*동$')


def parse_sheet(ws, dong_override=None):
    """단일 시트에서 벽체×층 인스턴스 추출. 동 표제는 시트 안에서 자동 검출하거나 override."""
    results = []
    current_dong = dong_override
    current_wall = None

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        row_text = ' '.join(str(c) if c is not None else '' for c in cells)

        # 동 표제 자동 검출 (override 있어도 갱신해줌 → 시트 내부 변경 대비)
        m_dong = DONG_HEADER_RE.search(row_text)
        if m_dong:
            current_dong = int(m_dong.group(1))
            continue

        if current_dong is None:
            continue

        # 벽 코드 (W1, CW1 등)
        for c in cells[:6]:
            if c is None:
                continue
            for line in str(c).strip().split('\n'):
                line = line.strip()
                m = re.match(r'^(\d+[A-Z]?)?-?([CHWD]+\d+[A-Z]?)$', line)
                if m and not m.group(0).startswith('('):
                    current_wall = m.group(2)
                    break

        # 층
        floor = None
        for c in cells:
            if c is None:
                continue
            s = str(c).strip()
            if re.match(r'^B?\d{1,2}F$', s):
                floor = s
                break

        # 두께
        thickness = None
        for c in cells:
            if c is None:
                continue
            try:
                v = int(c)
                if 100 <= v <= 600:
                    thickness = v
                    break
            except (ValueError, TypeError):
                pass

        # 배근
        rebars = []
        for c in cells:
            if c is None:
                continue
            m_re = re.findall(r'(D\d{1,2}@\d{2,3})', str(c).strip())
            rebars.extend(m_re)

        if current_wall and floor and thickness:
            results.append({
                'dong': current_dong,
                'wall_code': current_wall,
                'floor': floor,
                'thickness': thickness,
                'rebars': rebars,
            })

    return results


def parse_xlsx(xlsx_path):
    """xlsx 1개 → 시트별 추출 (BarList 단일 또는 시트명-동 다중)"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = wb.sheetnames
    all_results = []

    # 시트명이 동인 패턴
    sheet_dong_map = {}
    for sn in sheets:
        m = DONG_SHEET_RE.match(sn.strip())
        if m:
            sheet_dong_map[sn] = int(m.group(1))

    if sheet_dong_map:
        # 다중 시트 모드
        for sn, dong in sheet_dong_map.items():
            ws = wb[sn]
            all_results.extend(parse_sheet(ws, dong_override=dong))
    else:
        # 단일 BarList 모드
        ws = wb['BarList'] if 'BarList' in sheets else wb.active
        all_results.extend(parse_sheet(ws))

    return all_results


def clear_wall_members(territory_id):
    """기존 DONGxxx_WALL 행 일괄 삭제 — 재적재 위해."""
    import sqlite3
    con = sqlite3.connect(asset_db.DB_PATH)
    cur = con.cursor()
    cur.execute(
        "DELETE FROM members WHERE territory_id=? AND region LIKE 'DONG%_WALL'",
        (territory_id,),
    )
    deleted = cur.rowcount
    con.commit()
    con.close()
    return deleted


def main():
    asset_db.init_db()
    busan_id = asset_db.ensure_territory("부산 에코델타 24BL")

    # 정정 적재 위해 기존 WALL 행 삭제
    deleted = clear_wall_members(busan_id)
    print(f"[정정] 기존 DONGxxx_WALL 행 {deleted}건 삭제")

    total_walls = 0
    by_dong = Counter()
    unique_codes = {}
    parse_fail = 0

    for i, xlsx in enumerate(XLSX_FILES):
        if i % 100 == 0:
            print(f"  진행 {i}/{len(XLSX_FILES)}")
        try:
            rows = parse_xlsx(xlsx)
        except Exception as e:
            parse_fail += 1
            continue

        for r in rows:
            key = (r['dong'], r['wall_code'], r['thickness'])
            if key not in unique_codes:
                unique_codes[key] = r
            total_walls += 1
            by_dong[r['dong']] += 1

    print(f"\n[추출] 총 벽체×층 인스턴스: {total_walls}")
    print(f"[추출] 고유 (동, 코드, 두께) 조합: {len(unique_codes)}")
    print(f"[파싱 실패] {parse_fail}개 파일")
    print(f"[동별 분포 (인스턴스)]")
    for dong in sorted(by_dong.keys()):
        print(f"  {dong}동: {by_dong[dong]}건")

    # DB 적재
    inserted = 0
    insert_fail = 0
    by_dong_unique = Counter()
    for (dong, wall_code, thickness), r in unique_codes.items():
        region = f"DONG{dong}_WALL"
        rebars_str = ', '.join(r['rebars']) if r['rebars'] else None
        try:
            asset_db.insert_member(
                territory_id=busan_id,
                region=region,
                code=wall_code,
                member_type='WALL',
                thickness=thickness,
                main_bar=rebars_str,
            )
            inserted += 1
            by_dong_unique[dong] += 1
        except Exception:
            insert_fail += 1

    print(f"\n[DB 적재] 벽체 부재 {inserted}종 추가 (실패 {insert_fail})")
    print(f"[고유 부재 동별 분포]")
    for dong in sorted(by_dong_unique.keys()):
        print(f"  DONG{dong}_WALL: {by_dong_unique[dong]}종")

    print(f"\n=== 자산 DB 통계 ===")
    for table, cnt in asset_db.stats().items():
        print(f"  {table}: {cnt}")


if __name__ == '__main__':
    main()
