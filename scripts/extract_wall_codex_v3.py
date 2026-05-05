"""
벽체 일람표 OLE 추출 v3 — read_only + flush 강제판

v2가 메모리/버퍼 문제로 멈춘 후 재설계:
- read_only=True (openpyxl)
- 매 100개마다 stdout flush + 진척 출력
- 청크 처리 (50개씩 모아 DB 일괄 INSERT)
"""

import os, re, sys, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)

import openpyxl
import sqlite3
from collections import Counter
from core import asset_db

XLSX_FILES = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
print(f"[입력] xlsx {len(XLSX_FILES)}개", flush=True)

DONG_HEADER_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')
DONG_SHEET_RE = re.compile(r'^(\d{3})\s*동$')
WALL_CODE_RE = re.compile(r'^(\d+[A-Z]?)?-?([CHWD]+\d+[A-Z]?)$')
FLOOR_RE = re.compile(r'^B?\d{1,2}F$')
REBAR_RE = re.compile(r'(D\d{1,2}@\d{2,3})')


def parse_sheet(ws, dong_override=None):
    results = []
    current_dong = dong_override
    current_wall = None

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        # 동 표제 자동 검출
        for c in cells[:8]:
            if c is None:
                continue
            sval = str(c)
            m = DONG_HEADER_RE.search(sval)
            if m:
                current_dong = int(m.group(1))
                break

        if current_dong is None:
            continue

        # 벽 코드
        for c in cells[:6]:
            if c is None:
                continue
            for line in str(c).strip().split('\n'):
                line = line.strip()
                m = WALL_CODE_RE.match(line)
                if m and not m.group(0).startswith('('):
                    current_wall = m.group(2)
                    break

        # 층
        floor = None
        for c in cells:
            if c is None:
                continue
            s = str(c).strip()
            if FLOOR_RE.match(s):
                floor = s
                break

        # 두께
        thickness = None
        for c in cells:
            if c is None:
                continue
            try:
                v = int(c)
                if 100 <= v <= 600:
                    thickness = v
                    break
            except (ValueError, TypeError):
                pass

        # 배근
        rebars = []
        for c in cells:
            if c is None:
                continue
            rebars.extend(REBAR_RE.findall(str(c).strip()))

        if current_wall and floor and thickness:
            results.append({
                'dong': current_dong,
                'wall_code': current_wall,
                'floor': floor,
                'thickness': thickness,
                'rebars': rebars,
            })

    return results


def parse_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    all_results = []

    sheet_dong_map = {}
    for sn in sheets:
        m = DONG_SHEET_RE.match(sn.strip())
        if m:
            sheet_dong_map[sn] = int(m.group(1))

    if sheet_dong_map:
        for sn, dong in sheet_dong_map.items():
            ws = wb[sn]
            all_results.extend(parse_sheet(ws, dong_override=dong))
    else:
        ws = wb['BarList'] if 'BarList' in sheets else wb.active
        all_results.extend(parse_sheet(ws))

    wb.close()
    return all_results


def clear_wall_members(territory_id):
    con = sqlite3.connect(asset_db.DB_PATH)
    cur = con.cursor()
    cur.execute(
        "DELETE FROM members WHERE territory_id=? AND region LIKE 'DONG%_WALL'",
        (territory_id,),
    )
    deleted = cur.rowcount
    con.commit()
    con.close()
    return deleted


def main():
    asset_db.init_db()
    busan_id = asset_db.ensure_territory("부산 에코델타 24BL")

    deleted = clear_wall_members(busan_id)
    print(f"[정정] 기존 DONGxxx_WALL 행 {deleted}건 삭제", flush=True)

    total_walls = 0
    by_dong = Counter()
    unique_codes = {}
    parse_fail = 0

    for i, xlsx in enumerate(XLSX_FILES):
        if i % 50 == 0:
            print(f"  진행 {i}/{len(XLSX_FILES)}  unique={len(unique_codes)}  fail={parse_fail}", flush=True)
        try:
            rows = parse_xlsx(xlsx)
        except Exception as e:
            parse_fail += 1
            continue

        for r in rows:
            key = (r['dong'], r['wall_code'], r['thickness'])
            if key not in unique_codes:
                unique_codes[key] = r
            total_walls += 1
            by_dong[r['dong']] += 1

    print(f"\n[추출] 총 인스턴스: {total_walls}", flush=True)
    print(f"[추출] 고유 (동, 코드, 두께): {len(unique_codes)}", flush=True)
    print(f"[파싱 실패] {parse_fail}개", flush=True)
    print(f"[동별 인스턴스 분포]", flush=True)
    for dong in sorted(by_dong.keys()):
        print(f"  {dong}동: {by_dong[dong]}건", flush=True)

    # DB 적재 (단일 connection으로 일괄)
    con = sqlite3.connect(asset_db.DB_PATH)
    cur = con.cursor()
    inserted = 0
    insert_fail = 0
    by_dong_unique = Counter()
    for (dong, wall_code, thickness), r in unique_codes.items():
        region = f"DONG{dong}_WALL"
        rebars_str = ', '.join(r['rebars']) if r['rebars'] else None
        try:
            cur.execute(
                """INSERT OR IGNORE INTO members
                   (territory_id, region, code, member_type, thickness, main_bar)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (busan_id, region, wall_code, 'WALL', thickness, rebars_str),
            )
            if cur.rowcount > 0:
                inserted += 1
                by_dong_unique[dong] += 1
        except Exception:
            insert_fail += 1
    con.commit()
    con.close()

    print(f"\n[DB 적재] {inserted}종 추가 (실패 {insert_fail})", flush=True)
    print(f"[고유 부재 동별 분포]", flush=True)
    for dong in sorted(by_dong_unique.keys()):
        print(f"  DONG{dong}_WALL: {by_dong_unique[dong]}종", flush=True)

    print(f"\n=== 자산 DB 통계 ===", flush=True)
    for table, cnt in asset_db.stats().items():
        print(f"  {table}: {cnt}", flush=True)


if __name__ == '__main__':
    main()
