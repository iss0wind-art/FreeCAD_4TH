"""DB에 적재된 벽체 부재 5건을 무작위 추출 → 원본 xlsx와 대조"""
import sys, os, sqlite3, random, glob, re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
from core import asset_db

random.seed(7)

con = sqlite3.connect(asset_db.DB_PATH)
con.row_factory = sqlite3.Row
cur = con.cursor()
cur.execute("""
    SELECT region, code, thickness, main_bar
    FROM members
    WHERE region LIKE 'DONG%_WALL'
    ORDER BY RANDOM()
    LIMIT 5
""")
samples = cur.fetchall()

print("=== DB 샘플 5건 ===")
for s in samples:
    print(f"  {s['region']} | {s['code']} | t={s['thickness']} | bar={s['main_bar']}")

# 각 샘플 동에 해당하는 xlsx 무작위 1개에서 원본 패턴 검색
files = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
DONG_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')

print("\n=== 원본 대조 ===")
for s in samples:
    dong = int(s['region'].replace('DONG', '').replace('_WALL', ''))
    code = s['code']
    # 해당 동을 담은 xlsx 1개 찾기
    found_xlsx = None
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb['BarList'] if 'BarList' in wb.sheetnames else wb.active
            for row in ws.iter_rows(values_only=True, max_row=20):
                text = ' '.join(str(c) if c is not None else '' for c in row)
                m = DONG_RE.search(text)
                if m and int(m.group(1)) == dong:
                    found_xlsx = f
                    break
            wb.close()
            if found_xlsx:
                break
        except Exception:
            pass
    if not found_xlsx:
        print(f"  [{dong}동/{code}] 원본 xlsx 미검출")
        continue

    # 코드+두께 조합이 원본에 나오는지 확인
    wb = openpyxl.load_workbook(found_xlsx, data_only=True)
    ws = wb['BarList'] if 'BarList' in wb.sheetnames else wb.active
    code_seen = False
    thickness_seen_near_code = False
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        for c in row:
            if c is None:
                continue
            sval = str(c)
            if code in sval:
                code_seen = True
        if code_seen:
            # 같은 행에 두께 있나?
            for c in row:
                try:
                    if int(c) == s['thickness']:
                        thickness_seen_near_code = True
                except (ValueError, TypeError):
                    pass
    wb.close()
    status = "OK" if code_seen else "MISS"
    print(f"  [{dong}동/{code}/t={s['thickness']}] xlsx={os.path.basename(found_xlsx)} → 코드검출={code_seen}, 두께매칭={thickness_seen_near_code} [{status}]")
