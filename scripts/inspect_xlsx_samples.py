"""xlsx 10개 샘플 검수 — BarList 구조, 동 표제 패턴, 데이터 분량 확인"""
import os, sys, glob, random, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

random.seed(42)
files = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
samples = random.sample(files, 10)

DONG_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')
DONG_RE_LOOSE = re.compile(r'(\d{3})\s*동')

for path in samples:
    name = os.path.basename(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[FAIL] {name}: {e}")
        continue

    sheets = wb.sheetnames
    ws = wb['BarList'] if 'BarList' in sheets else wb.active

    rows = list(ws.iter_rows(values_only=True))
    nrows = len(rows)

    # 동 표제 검출 (엄격/관대)
    strict_hits = []
    loose_hits = []
    for i, row in enumerate(rows):
        text = ' '.join(str(c) if c is not None else '' for c in row)
        for m in DONG_RE.finditer(text):
            strict_hits.append((i, m.group(1)))
        for m in DONG_RE_LOOSE.finditer(text):
            loose_hits.append((i, m.group(1)))

    # 첫 5행 표시 (구조 파악)
    print(f"\n=== {name} ===")
    print(f"  시트: {sheets}, 사용 시트: {ws.title}, 총 행수: {nrows}")
    print(f"  엄격 <NNN동> 검출: {len(strict_hits)} → {strict_hits[:5]}")
    print(f"  관대 NNN동 검출: {len(loose_hits)} → {loose_hits[:5]}")

    print(f"  [상위 8행]")
    for i, row in enumerate(rows[:8]):
        cells = [str(c)[:20] if c is not None else '' for c in row[:10]]
        print(f"    {i:3d}: {cells}")
