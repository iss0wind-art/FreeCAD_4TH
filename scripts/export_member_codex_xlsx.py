"""
통합 부재 법전 → Excel 보고서 생성

방부장께 보고용. 부재 종류별 시트 + 미분류 별도 시트.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from core import asset_db

OUT_PATH = r"D:/06.3지국 전용방/통합부재법전_2026-05-05.xlsx"

# 모든 부재 조회
members = asset_db.query("""
    SELECT m.region, m.code, m.member_type,
           m.floor_from, m.floor_to,
           m.width, m.height, m.thickness,
           m.main_bar, m.hoop,
           m.extracted_at
    FROM members m
    ORDER BY m.member_type, m.region, m.code
""")

print(f"[조회] 총 {len(members)}종")

# 1차 목표 = 거푸집·콘크리트 물량 산출 (방부장 친명, 2026-05-05)
# 철근 정보(주근·후프·배근)는 차후 업데이트 사안 — 분류 등급에서 무시
def classify(m):
    mt = m['member_type']
    if mt == 'COLUMN':
        if not m['width'] or not m['height']: return 'UNCLASSIFIED', '단면 누락'
        return 'COMPLETE', ''
    elif mt == 'BEAM':
        if not m['width'] or not m['height']: return 'UNCLASSIFIED', '단면 누락'
        return 'COMPLETE', ''
    elif mt == 'WALL':
        if not m['thickness']: return 'UNCLASSIFIED', '두께 누락'
        return 'COMPLETE', ''
    elif mt == 'SLAB':
        if not m['thickness']: return 'UNCLASSIFIED', '두께 누락'
        return 'COMPLETE', ''
    return 'UNCLASSIFIED', f'알 수 없는 부재종류: {mt}'

complete, partial, unclassified = [], [], []
for m in members:
    grade, issue = classify(m)
    m_with = {**m, 'issues': issue, 'grade': grade}
    if grade == 'COMPLETE':
        complete.append(m_with)
    elif grade == 'PARTIAL':
        partial.append(m_with)
    else:
        unclassified.append(m_with)

print(f"[분류] 완전 {len(complete)}종 / 부분 {len(partial)}종 / 미분류 {len(unclassified)}종")

# Excel 생성
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

def write_sheet(ws, rows, title):
    headers = ['영역', '코드', '부재종류', '층 시작', '층 끝',
               '폭(mm)', '높이(mm)', '두께(mm)', '주근', '후프', '비고']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for i, m in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=m['region'])
        ws.cell(row=i, column=2, value=m['code'])
        ws.cell(row=i, column=3, value=m['member_type'])
        ws.cell(row=i, column=4, value=m['floor_from'])
        ws.cell(row=i, column=5, value=m['floor_to'])
        ws.cell(row=i, column=6, value=m['width'])
        ws.cell(row=i, column=7, value=m['height'])
        ws.cell(row=i, column=8, value=m['thickness'])
        ws.cell(row=i, column=9, value=m['main_bar'])
        ws.cell(row=i, column=10, value=m['hoop'])
        ws.cell(row=i, column=11, value=m.get('issues', ''))
        if i % 2 == 0:
            for c in range(1, 12):
                ws.cell(row=i, column=c).fill = ALT_FILL
        if m.get('issues'):
            ws.cell(row=i, column=11).fill = WARN_FILL
    # 컬럼 폭
    widths = [18, 16, 12, 8, 8, 10, 10, 10, 18, 14, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

# 시트 1 — 요약
ws_summary = wb.active
ws_summary.title = "00_요약"
ws_summary['A1'] = "통합 부재 법전 — 부산 에코델타 24BL"
ws_summary['A1'].font = Font(size=14, bold=True)
ws_summary['A2'] = f"산출일시: {datetime.now().isoformat(timespec='seconds')}"
ws_summary['A4'] = "분류"
ws_summary['B4'] = "수량(종)"
ws_summary['A4'].font = HEADER_FONT; ws_summary['A4'].fill = HEADER_FILL
ws_summary['B4'].font = HEADER_FONT; ws_summary['B4'].fill = HEADER_FILL

stats = asset_db.query("""
    SELECT member_type, region, COUNT(*) AS n
    FROM members GROUP BY member_type, region ORDER BY member_type, region
""")
row = 5
for s in stats:
    ws_summary.cell(row=row, column=1, value=f"{s['member_type']} - {s['region']}")
    ws_summary.cell(row=row, column=2, value=s['n'])
    row += 1
ws_summary.cell(row=row, column=1, value="합계")
ws_summary.cell(row=row, column=2, value=len(members))
ws_summary.cell(row=row, column=1).font = Font(bold=True)
ws_summary.cell(row=row, column=2).font = Font(bold=True)
row += 2
ws_summary.cell(row=row, column=1, value="완전 (모든 데이터)").font = Font(bold=True)
ws_summary.cell(row=row, column=2, value=len(complete))
ws_summary.cell(row=row+1, column=1, value="부분 (배근/후프 누락)")
ws_summary.cell(row=row+1, column=2, value=len(partial))
ws_summary.cell(row=row+2, column=1, value="미분류 (단면/두께 누락)")
ws_summary.cell(row=row+2, column=2, value=len(unclassified))
ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 12

# 시트 2 — 기둥 (영역별)
ws_col = wb.create_sheet("01_기둥")
write_sheet(ws_col, [m for m in complete if m['member_type'] == 'COLUMN'], "기둥")

# 시트 3 — 보
ws_beam = wb.create_sheet("02_보")
write_sheet(ws_beam, [m for m in complete if m['member_type'] == 'BEAM'], "보")

# 시트 4 — 벽체 (동별)
ws_wall = wb.create_sheet("03_벽체")
write_sheet(ws_wall, [m for m in complete if m['member_type'] == 'WALL'], "벽체")

# 시트 5 — 슬라브
ws_slab = wb.create_sheet("04_슬라브")
write_sheet(ws_slab, [m for m in complete if m['member_type'] == 'SLAB'], "슬라브")

# 시트 6 — 부분 (1차에서 제외, 차후 철근 보강 대상 — 비어있어야 정상)
if partial:
    ws_partial = wb.create_sheet("98_부분")
    write_sheet(ws_partial, partial, "부분")

# 시트 7 — 미분류
if unclassified:
    ws_un = wb.create_sheet("99_미분류")
    write_sheet(ws_un, unclassified, "미분류")

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"[저장] {OUT_PATH}")
print(f"  파일 크기: {os.path.getsize(OUT_PATH)/1024:.1f} KB")
