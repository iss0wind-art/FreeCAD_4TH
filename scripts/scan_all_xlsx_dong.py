"""전체 981 xlsx 스캔 — 동 표제·시트명 분포 집계"""
import os, sys, glob, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
from collections import Counter, defaultdict

DONG_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')
DONG_SHEET_RE = re.compile(r'^(\d{3})\s*동$')

files = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
print(f"[입력] xlsx {len(files)}개")

dong_counter = Counter()
multi_sheet_files = []
no_dong_files = []
multi_dong_in_one_xlsx = []  # 한 xlsx에 동 표제 2개 이상

for i, path in enumerate(files):
    if i % 100 == 0:
        print(f"  진행 {i}/{len(files)}")
    name = os.path.basename(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        no_dong_files.append((name, f"OPEN_FAIL: {e}"))
        continue

    sheets = wb.sheetnames
    dong_set = set()

    # 시트명에서 동 번호 추출
    sheet_dongs = []
    for sn in sheets:
        m = DONG_SHEET_RE.match(sn.strip())
        if m:
            sheet_dongs.append(int(m.group(1)))
            dong_set.add(int(m.group(1)))

    if len(sheet_dongs) >= 1:
        # 시트명이 동인 경우 — 각 시트가 독립 동
        multi_sheet_files.append((name, sheet_dongs))
        for d in sheet_dongs:
            dong_counter[d] += 1
        wb.close()
        continue

    # BarList 또는 첫 시트에서 <NNN동> 검출
    ws = wb['BarList'] if 'BarList' in sheets else wb.active
    found = []
    for row in ws.iter_rows(values_only=True, max_row=20):
        text = ' '.join(str(c) if c is not None else '' for c in row)
        for m in DONG_RE.finditer(text):
            found.append(int(m.group(1)))

    if not found:
        no_dong_files.append((name, "NO_DONG_HEADER"))
    else:
        unique_d = list(dict.fromkeys(found))
        if len(unique_d) > 1:
            multi_dong_in_one_xlsx.append((name, unique_d))
        for d in unique_d:
            dong_counter[d] += 1
            dong_set.add(d)
    wb.close()

print(f"\n=== 동별 xlsx 수 ===")
for d in sorted(dong_counter.keys()):
    print(f"  {d}동: {dong_counter[d]}개")

print(f"\n=== 시트명이 동인 파일 (multi-sheet) ===")
print(f"  총 {len(multi_sheet_files)}개")
for name, dongs in multi_sheet_files[:10]:
    print(f"  {name}: {dongs}")

print(f"\n=== 동 표제 미검출 파일 ===")
print(f"  총 {len(no_dong_files)}개")
for name, reason in no_dong_files[:10]:
    print(f"  {name}: {reason}")

print(f"\n=== 한 xlsx에 동 표제 2개 이상 ===")
print(f"  총 {len(multi_dong_in_one_xlsx)}개")
for name, dongs in multi_dong_in_one_xlsx[:10]:
    print(f"  {name}: {dongs}")
