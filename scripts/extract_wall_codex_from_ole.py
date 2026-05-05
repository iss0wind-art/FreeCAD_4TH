"""
벽체 일람표 OLE 추출 → 부재 법전 적재

방부장 친명 (2026-05-05): 단군과 합세하여 진행. 도면 전문가 검증 완료된
OLE 추출 방식을 일괄 적용.

소스: tmp_analysis/ole_extracted/*.xlsx (3개 도면, 16동 분량)
적재: jiguk3_assets.db members 테이블 (region=DONG{nnn}_WALL)
"""

import os, re, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from core import asset_db

import glob
XLSX_FILES = sorted(glob.glob(r"d:/Git/FreeCAD_4TH/tmp_analysis/ole_all/*.xlsx"))
print(f"[입력] xlsx {len(XLSX_FILES)}개")

DONG_HEADER_RE = re.compile(r'<\s*(\d{3})\s*동\s*>')
WALL_CODE_RE = re.compile(r'^([A-Z]+\d+[A-Z]?)')

def parse_xlsx(xlsx_path):
    """xlsx에서 동별 벽체 일람 추출"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['BarList'] if 'BarList' in wb.sheetnames else wb.active

    results = []  # (dong, wall_code, floor, thickness, vert_rebar, horiz_rebar, edge_rebar)
    current_dong = None
    current_wall = None
    current_thickness = None
    current_vert = None
    current_horiz = None
    current_edge = None

    for row in ws.iter_rows(values_only=True):
        # 모든 셀 텍스트 합치기 (동 표제 검출용)
        row_text = ' '.join(str(c) if c else '' for c in row)

        # 동 표제 (<101동>)
        m_dong = DONG_HEADER_RE.search(row_text)
        if m_dong:
            current_dong = int(m_dong.group(1))
            continue

        if current_dong is None:
            continue

        # 셀별 직접 분석 — 벽 코드, 층, 두께, 배근 위치 식별
        cells = list(row)

        # 벽 코드 후보 (WALL 컬럼: 보통 col index 2 부근, "59A-W1\n(토압 받는 벽체)" 같은 형식)
        for c in cells[:6]:
            if c is None: continue
            s = str(c).strip()
            # 벽 코드 패턴: W1, CW1, HW1, DW1 등 또는 59A-W1 같은 형식
            for line in s.split('\n'):
                line = line.strip()
                m = re.match(r'^(\d+[A-Z]?)?-?([CHWD]+\d+[A-Z]?)$', line)
                if m and not m.group(0).startswith('('):
                    current_wall = m.group(2)
                    break

        # 층 정보 (floor) — "16F", "B1F" 등
        floor = None
        for c in cells:
            if c is None: continue
            s = str(c).strip()
            if re.match(r'^B?\d{1,2}F$', s):
                floor = s
                break

        # 두께 (mm) — 100~600 범위 정수
        thickness = None
        for c in cells:
            if c is None: continue
            try:
                v = int(c)
                if 100 <= v <= 600:
                    thickness = v
                    break
            except (ValueError, TypeError):
                pass

        # 배근 패턴 — D10@250, D13@200 등
        rebars = []
        for c in cells:
            if c is None: continue
            s = str(c).strip()
            m = re.findall(r'(D\d{1,2}@\d{2,3})', s)
            rebars.extend(m)

        # 데이터 한 행 = 한 벽체×층 인스턴스
        if current_wall and floor and thickness:
            results.append({
                'dong': current_dong,
                'wall_code': current_wall,
                'floor': floor,
                'thickness': thickness,
                'rebars': rebars,
            })

    return results

def floor_to_int(floor_str):
    """16F → 16, B1F → -1, B2F → -2"""
    if floor_str.startswith('B'):
        return -int(floor_str[1:-1])
    return int(floor_str[:-1])

def main():
    asset_db.init_db()
    busan_id = asset_db.ensure_territory("부산 에코델타 24BL")

    total_walls = 0
    by_dong = {}
    unique_codes = {}  # (dong, wall_code, thickness) → 첫 번째 발견

    for xlsx in XLSX_FILES:
        if not os.path.exists(xlsx):
            print(f"[건너뜀] {xlsx}")
            continue
        print(f"[처리] {os.path.basename(xlsx)}")
        rows = parse_xlsx(xlsx)
        print(f"  추출 행수: {len(rows)}")

        for r in rows:
            key = (r['dong'], r['wall_code'], r['thickness'])
            if key not in unique_codes:
                unique_codes[key] = r
            total_walls += 1
            by_dong.setdefault(r['dong'], 0)
            by_dong[r['dong']] += 1

    print(f"\n[추출] 총 벽체×층 인스턴스: {total_walls}")
    print(f"[추출] 고유 (동, 코드, 두께) 조합: {len(unique_codes)}")
    print(f"[동별 분포]")
    for dong in sorted(by_dong.keys()):
        print(f"  {dong}동: {by_dong[dong]}건")

    # DB 적재
    inserted = 0
    for (dong, wall_code, thickness), r in unique_codes.items():
        region = f"DONG{dong}_WALL"
        rebars_str = ', '.join(r['rebars']) if r['rebars'] else None
        try:
            asset_db.insert_member(
                territory_id=busan_id,
                region=region,
                code=wall_code,
                member_type='WALL',
                thickness=thickness,
                main_bar=rebars_str,
            )
            inserted += 1
        except Exception as ex:
            pass

    print(f"\n[DB 적재] 벽체 부재 {inserted}종 추가")

    # 갱신된 통계
    print(f"\n=== 자산 DB 통계 ===")
    for table, cnt in asset_db.stats().items():
        print(f"  {table}: {cnt}")

if __name__ == '__main__':
    main()
